import pygsheets
import openpyxl
import datetime
import dateutil
from dateutil import parser
from datetime import date, timedelta
from openpyxl import Workbook
ouf = open ('file.txt', 'w')
gc = pygsheets.authorize(outh_file='client_secretxxx.json')
sh=gc.open('Сегафредо')
wks=sh[5]


wb = openpyxl.load_workbook('Сегафредо.xlsx')
#print(wb.get_sheet_names())
sheet = wb['Ашан Регионы']
sheet2 = wb['Окей Регионы']

if date.today().weekday()==0: #определяем нужную среду
    wednesday = date.today() - timedelta(5)
elif date.today().weekday()==1:
    wednesday = date.today() - timedelta(6)
elif date.today().weekday()==2:
    wednesday = date.today() - timedelta(7)
elif date.today().weekday()==3:
    wednesday = date.today() - timedelta(1)
elif date.today().weekday()==4:
    wednesday = date.today() - timedelta(2)
elif date.today().weekday()==5:
    wednesday = date.today() - timedelta(3)
elif date.today().weekday()==6:
    wednesday = date.today() - timedelta(4)

cities = [] # список городов

for city in range(3,31):
    cities+=[sheet.cell(row=city, column=1).value]
#print (cities)

stores = [] #список магазинов

for store in range (3,31):
    stores +=[sheet.cell(row=store, column=2).value]
#print (stores)

dates = []
for date in range (3,31):
    dates +=[sheet.cell(row=date, column=6).value.date()]
#print(dates)
positions = [] # список товаров
for position in range (7,43):
    positions +=[sheet.cell(row = 2, column = position).value]
#print (positions)
net_fotok=[]
for count in range (2,30):
    netu=[]
    zeroes=[]
    count+=1
    print (cities[count-3], end='')
    ouf.write('***  ' + cities[count-3] + ' ')
    print (stores[count-3])
    ouf.write(stores[count-3] + '   ***' + '\n')
    if dates[count-3] < wednesday + timedelta(1):
        print ('Нет фотографий')
        ouf.write('Нет фотографий' + '\n')
        net_fotok+=[str(cities[count-3]) +' '+ str(stores[count-3])]
        continue
    for zero in range (7,43):
        zeroes +=[sheet.cell(row=count, column=zero).value]
    #print (zeroes)
    for index, i in enumerate (zeroes):
        if i==None or i==0:
            print (positions[index])
            ouf.write(positions[index] + '\n')
            netu+=[positions[index]]
            sheet.cell(row=count-1, column =52, value=str(netu[:]))
cities2 = [] # список городов

for city2 in range(3,31):
    cities2+=[sheet2.cell(row=city2, column=1).value]
#print (cities)

stores2 = [] #список магазинов

for store2 in range (3,7):
    stores2 +=[sheet2.cell(row=store2, column=2).value]

dates2 = []
for date in range (3,7):
    dates2 +=[sheet2.cell(row=date, column=7).value.date()]
#print (dates2)
positions2 = []
for position2 in range (7,37):
    positions2 +=[sheet2.cell(row = 2, column = position2).value]
#print (positions2)
for count2 in range (2,6):
    netu2 =[]
    zeroes2=[]
    count2+=1
    print (cities2[count2-3], end=' ')
    ouf.write ('***   ' + cities2[count2-3]+ ' ')
    print (stores2[count2-3])
    ouf.write(stores2[count2-3] + '   ***' + '\n')
    if dates[count2-3] < wednesday + timedelta(1):
        print ('Нет фотографий')
        ouf.write('Нет фотографий' + '\n')
        net_fotok+=[str(cities2[count2-3]) +' '+ str(stores2[count2-3])]
        continue
    for zero2 in range (7,37):
        zeroes2 +=[sheet2.cell(row=count2, column=zero2).value]
    #print (zeroes)
    for index, j in enumerate (zeroes2):
        if j==None or j==0:
            print (positions2[index])
            ouf.write(positions2[index] + '\n')
            netu2+=[positions2[index]]
            sheet2.cell(row=count2, column =45, value=str(netu2[:]))
print ('Нет фотографий по:', str(net_fotok[:]))
ouf.write('\n' + 'Нет фотографий по:' + str(net_fotok[:]))

#wb.save('Сегафредо1.xlsx')
#sheet.cell(coordinate="C3").value=3
#print (sheet['AZ20'].value)
wks.clear('BA3', 'BA30')


for count in range (3,31):
    wks.update_cell('BA'+str(count), sheet['AZ'+str(count-1)].value)
wks2=sh[6]
wks2.clear('AS3', 'AS6')
count=0
for count2 in range (3,7):
    wks2.update_cell('AS'+str(count2), sheet2['AS'+str(count2)].value)
