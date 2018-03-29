import pygsheets
import openpyxl
gc = pygsheets.authorize(outh_file='client_secretxxx.json')
sh=gc.open('Copy of Сегафредо')
wks=sh[5]

wb = openpyxl.load_workbook('./Сегафредо.xlsx')
#print(wb.get_sheet_names())
sheet = wb['Ашан Регионы']



cities = [] # список городов

for city in range(3,31):
    cities+=[sheet.cell(row=city, column=1).value]
#print (cities)

stores = [] #список магазинов

for store in range (3,31):
    stores +=[sheet.cell(row=store, column=2).value]
#print (stores)

positions = [] # список товаров
for position in range (7,43):
    positions +=[sheet.cell(row = 2, column = position).value]
#print (positions)

for count in range (3,30):
    netu=[]
    zeroes=[]
    count+=1
    print (cities[count-3], end='')
    print (stores[count-3])
    for zero in range (7,43):
        zeroes +=[sheet.cell(row=count, column=zero).value]
    #print (zeroes)
    for index, i in enumerate (zeroes):
        if i==None or i=='0':
            print (positions[index])
            netu+=[positions[index]]
            sheet['AZ'+str(count-1)]=str(netu[:])

wks.clear('AY3', 'AY30')

for count in range (3,30):
    wks.update_cell('AY'+str(count), sheet['AZ'+str(count)])
